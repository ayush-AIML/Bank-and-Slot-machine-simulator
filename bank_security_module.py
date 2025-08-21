import openpyxl as xl

bank_workbook = xl.load_workbook("bank_data.xlsx")
data_sheet = bank_workbook['Sheet1']

# We are going to store the usernames and passwords here
usernames = []
passwords = []

for i in range(2, data_sheet.max_row + 1):
    usernames.append(data_sheet.cell(i, 1).value)
    passwords.append(data_sheet.cell(i, 2).value)


def new_user():
    name = str(input("Enter your name: "))
    while True:
        new_username = str(input("Enter username: "))
        if new_username in usernames:
            print("Username already taken...")
            continue
        else:
            break
    new_password = str(input("Enter your password: "))
    data_sheet.cell(data_sheet.max_row + 1, 1).value = new_username
    data_sheet.cell(data_sheet.max_row, 2).value = new_password
    data_sheet.cell(data_sheet.max_row, 3).value = 0
    data_sheet.cell(data_sheet.max_row, 4).value = name
    bank_workbook.save("bank_data.xlsx")
    print("Account created Successfully!")


def uname_check():
    while True:
        u = input("Enter username: ")
        if u in usernames:
            login_credential = usernames.index(u) + 2  # Adding 2 to align the login credential to the user row value
            name = data_sheet.cell(login_credential, 4).value
            return name, login_credential
        else:
            print("No user found")
            continue


def pword_check(login_value):
    while True:
        p = input("Enter password: ")
        if p == data_sheet.cell(login_value, 2).value:
            return p
        else:
            print("Wrong password")
            continue


def account_info(login_credential):
    balance = data_sheet.cell(login_credential, 3).value
    return balance


def money_deposit(deposit_credential, amount):
    deposit_cell = data_sheet.cell(deposit_credential, 3)
    deposit_cell.value += amount
    bank_workbook.save("bank_data.xlsx")


def money_withdraw(withdraw_credential, amount):
    withdraw_cell = data_sheet.cell(withdraw_credential, 3)
    withdraw_cell.value -= amount
    bank_workbook.save("bank_data.xlsx")
